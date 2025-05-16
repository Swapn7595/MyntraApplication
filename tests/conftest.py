# conftest.py
import os
import pytest
import logging
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from utils.config import environments
from utils.config import logger
from utils.config import test_step_logger
from utils.config import TestStepLogger
import openpyxl
from openpyxl.styles import Font

# Create or load Excel workbook
report_path = "reports/test_report.xlsx"
if os.path.exists(report_path):
    workbook = openpyxl.load_workbook(report_path)
    sheet = workbook.active
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Test Report"
    headers = ["Test Case ID", "Test Title", "Test Steps", "Status", "Error Message", "Execution Time"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

# Track current row for new data
current_row = sheet.max_row + 1

# Initialize TestStepLogger
test_step_logger = TestStepLogger()

# Track existing test case IDs to avoid duplicates
existing_test_ids = set((row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row))

# Counter for test case IDs
test_case_counter = sheet.max_row - 1 if sheet.max_row > 1 else 0


def pytest_addoption(parser):
    parser.addoption(
        "--env", action="store", default="dev", help="Environment to run tests against"
    )


@pytest.fixture(scope='session')
def config(request):
    env = request.config.getoption("--env")
    if env in environments:
        return environments[env]
    else:
        raise ValueError(f"Unknown environment: {env}")


@pytest.fixture(scope='session')
def driver(config):
    chrome_options = Options()
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--log-level=3')

    service = Service()
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.maximize_window()

    base_url = config['base_url']
    driver.get(base_url)

    yield driver
    driver.quit()


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_protocol(item, nextitem):
    global test_case_counter
    start_time = datetime.now()
    test_step_logger.steps = []
    yield
    end_time = datetime.now()

    test_name = item.name
    if "::" in test_name:
        test_name = test_name.rsplit("::", 1)[-1]

    test_title = " ".join(word.capitalize() for word in test_name.split("_"))
    test_case_counter += 1
    test_case_id = f"TC_{test_case_counter:03d}"
    test_steps = test_step_logger.get_steps() or "N/A"

    item.user_properties.append(("execution_time", str(end_time - start_time)))
    item.user_properties.append(("test_case_id", test_case_id))
    item.user_properties.append(("test_title", test_title))
    item.user_properties.append(("test_steps", test_steps))


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    if report.when == "call":
        status = "Passed" if report.outcome == "passed" else "Failed"
        error_message = "" if status == "Passed" else str(report.longrepr) if hasattr(report, 'longrepr') else "Unknown Error"

        execution_time = dict(item.user_properties).get("execution_time", "N/A")
        test_case_id = dict(item.user_properties).get("test_case_id", "Unknown")
        test_title = dict(item.user_properties).get("test_title", "Unknown")
        test_steps = dict(item.user_properties).get("test_steps", "N/A")

        global current_row
        if (test_case_id, test_title) not in existing_test_ids:
            sheet.cell(row=current_row, column=1, value=test_case_id)
            sheet.cell(row=current_row, column=2, value=test_title)
            sheet.cell(row=current_row, column=3, value=test_steps)
            sheet.cell(row=current_row, column=4, value=status)
            sheet.cell(row=current_row, column=5, value=error_message)
            sheet.cell(row=current_row, column=6, value=execution_time)
            existing_test_ids.add((test_case_id, test_title))
            current_row += 1

    if report.failed:
        driver = item.funcargs.get("driver")
        if driver:
            screenshots_dir = os.path.join(os.getcwd(), ".screenshots")
            os.makedirs(screenshots_dir, exist_ok=True)

            test_name = item.nodeid.replace("::", "_").replace("/", "_")
            exc_type = type(call.excinfo.value).__name__ if call.excinfo else "Exception"
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            screenshot_filename = f"{test_name}_{exc_type}_{timestamp}.png"
            screenshot_path = os.path.join(screenshots_dir, screenshot_filename)

            driver.save_screenshot(screenshot_path)
            print(f"\n[SCREENSHOT SAVED] {screenshot_path}")


@pytest.hookimpl(trylast=True)
def pytest_sessionfinish(session, exitstatus):
    try:
        workbook.save(report_path)
        print(f"Enhanced test report saved to {report_path}")
    except Exception as e:
        print(f"Failed to save test report: {e}")
